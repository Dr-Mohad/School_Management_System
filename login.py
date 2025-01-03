import pandas as pd
import datetime
import time
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
def register_students():
    # Load the data
    main_data = pd.read_excel('Registeration.xlsx', dtype={'Student_Number': str, 'Guardian_Number': str})
    i_s = main_data.ID.tolist()
    n = main_data.Student_Number.tolist()
    n2 = main_data.Guardian_Number.tolist()
    continue_registration = True

    # Function to generate a new unique ID
    def generate_next_id(existing_ids):
        if not existing_ids:
            return 'MS001'
        max_id = max(int(i[2:]) for i in existing_ids if i.startswith('MS'))
        return f"MS{max_id + 1:03}"

    while continue_registration:
        try:
            # Ask how many students to register
            num_students = int(input('How many students would you like to register? '))
        except ValueError:
            print('Please enter a valid number.')
            continue

        # Loop through the number of students to register
        for _ in range(num_students):
            # Automatically generate a new ID
            iD = generate_next_id(i_s)

            print(f"The new ID for the student is: {iD}")

            # Validate Full Name
            while True:
                full_name = input('Enter Full Name (First Middle Last): ').title().strip().split()
                if all(len(i) >= 4 for i in full_name) and len(full_name) == 3:
                    break
                else:
                    print("The full name is invalid. Ensure it is three words, each at least 4 characters long.")

            # Validate Student Number
            while True:
                student_number = input('Enter Student Number: ')
                if student_number in n or student_number in n2:
                    print('This number is already registered. Please try again.')
                elif student_number.startswith('061') and len(student_number) == 10 and student_number.isdigit():
                    break
                else:
                    print("Invalid student number. Please ensure it starts with '061' and is 10 digits long.")

            # Validate Gender
            while True:
                gender = input('Enter Gender (Male/Female): ').capitalize()
                if gender in ['Male', 'Female']:
                    break
                else:
                    print("Please enter either Male or Female.")

            # Validate Guardian Name
            while True:
                guardian_name = input('Enter Guardian Name (Full Name): ')
                if all(len(part) >= 4 for part in guardian_name.split()) and len(guardian_name.split()) == 3:
                    break
                else:
                    print("Please enter a valid guardian full name with three words.")

            # Validate Guardian Number
            while True:
                guardian_number = input('Enter Guardian Number: ')
                if guardian_number in n2 or guardian_number in n:
                    print('This number is already registered. Please try again.')
                elif guardian_number.startswith('061') and len(guardian_number) == 10 and guardian_number.isdigit():
                    break
                else:
                    print("Invalid guardian number. Please ensure it starts with '061' and is 10 digits long.")

            # Validate Date of Birth
            while True:
                dob = input('Enter Date of Birth (dd/mm/yyyy): ')
                try:
                    day, month, year = map(int, dob.split('/'))
                    current_year = datetime.datetime.now().year
                    if 1 <= day <= 31 and 1 <= month <= 12 and year <= (current_year - 7):
                        break
                    else:
                        print("Invalid date of birth. Please try again.")
                except ValueError:
                    print("Please use the correct format: dd/mm/yyyy.")

            # Validate Class
            while True:
                print('1. Form One\n2. Form Two\n3. Form Three\n4. Form Four')
                class_choice = input('Enter Class: ')
                if class_choice == '1':
                    class_name = 'Form One'
                    sheet_name = 'Form1'
                    attendance_data = pd.read_excel('Attendence.xlsx', sheet_name=sheet_name)
                    exam_data = pd.read_excel('Exam result.xlsx', sheet_name=sheet_name)
                    break
                elif class_choice == '2':
                    class_name = 'Form Two'
                    sheet_name = 'Form2'
                    attendance_data = pd.read_excel('Attendence.xlsx', sheet_name=sheet_name)
                    exam_data = pd.read_excel('Exam result.xlsx', sheet_name=sheet_name)
                    break
                elif class_choice == '3':
                    class_name = 'Form Three'
                    sheet_name = 'Form3'
                    attendance_data = pd.read_excel('Attendence.xlsx', sheet_name=sheet_name)
                    exam_data = pd.read_excel('Exam result.xlsx', sheet_name=sheet_name)
                    break
                elif class_choice == '4':
                    class_name = 'Form Four'
                    sheet_name = 'Form4'
                    attendance_data = pd.read_excel('Attendence.xlsx', sheet_name=sheet_name)
                    exam_data = pd.read_excel('Exam result.xlsx', sheet_name=sheet_name)
                    break
                else:
                    print('Please enter a valid number.')

            # Create a new student record
            new_student = {
                'ID': iD,
                'Full_Name': " ".join(full_name),
                'Student_Number': student_number,
                'Gender': gender,
                'Guardian_Name': guardian_name,
                'Guardian_Number': guardian_number,
                'Date_of_Birth': dob,
                'Registration_Date': time.strftime('%d/%m/%Y'),
                'Class': class_name
            }
            # Create a new attendance record
            attendance_record = {
                'ID': iD,
                'Full_Name': " ".join(full_name),
                'Absent': 0
            }
            # Create a new exam record
            exam_record = {
                'ID_Number': iD,
                'Full_Name': " ".join(full_name),
                'Monthly_Exam_1': 0,
                'Mid_Term': 0,
                'Monthly_Exam_2': 0,
                'Final': 0,
                'Total': 0
            }

            # Save the new records
            exam_data = pd.concat([exam_data, pd.DataFrame([exam_record])], ignore_index=True)
            with pd.ExcelWriter('Exam result.xlsx', mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                exam_data.to_excel(writer, sheet_name=sheet_name, index=False)

            attendance_data = pd.concat([attendance_data, pd.DataFrame([attendance_record])], ignore_index=True)
            with pd.ExcelWriter('Attendence.xlsx', mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                attendance_data.to_excel(writer, sheet_name=sheet_name, index=False)

            main_data = pd.concat([main_data, pd.DataFrame([new_student])], ignore_index=True)
            with pd.ExcelWriter('Registeration.xlsx', mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                main_data.to_excel(writer, index=False)

            i_s.append(iD)
            n.append(student_number)
            n2.append(guardian_number)

        while True:
            continue_prompt = input('Do you want to continue registration? [1. Yes / 2. No] ')
            if continue_prompt == '1':
                break
            elif continue_prompt == '2':
                continue_registration = False
                break
            else:
                print('Invalid input. Please try again.')

    print("Registration completed successfully.")
def Xog_raadin():
    # Load data from Excel files
    registration_data = pd.read_excel('Registeration.xlsx', dtype={'Number_ka': str, 'Masuul_no': str})
    exam_data = pd.read_excel('Exam result.xlsx')
    
    # Extract the necessary columns into lists
    student_ids = registration_data.ID.tolist()
    full_names = registration_data.Full_Name.tolist()
    phone_numbers = registration_data.Student_Number.tolist()
    genders = registration_data.Gender.tolist()
    guardian_names = registration_data.Guardian_Name.tolist()
    guardian_numbers = registration_data.Guardian_Number.tolist()
    birth_dates = registration_data.Date_of_Birth.tolist()
    registration_dates = registration_data.Registration_Date.tolist()
    classes = registration_data.Class.tolist()

    monthly_exam_1 = exam_data.Monthly_Exam_1.tolist()
    mid_term = exam_data.Mid_term.tolist()
    monthly_exam_2 = exam_data.Monthly_Exam_2.tolist()
    final_exam = exam_data.Final.tolist()
    total_marks = exam_data.Total.tolist()

    # Start the loop for student search
    c_s = True
    while c_s:
        print('')
        student_id = input('Enter  ID Student: ') 
        if student_id in student_ids:
            index = student_ids.index(student_id)  # Find the index of the ID

            # Display student information
            print("\n--- Student Information ---")
            print('')
            print(f"{'Field':<20}{'Value':<30}")
            print(f"{'-'*50}")
            print(f"{'Student ID':<20}{student_id:<30}")
            print(f"{'Full_Name':<20}{full_names[index]:<30}")
            print(f"{'Student_Number':<20}{phone_numbers[index]:<30}")
            print(f"{'Gender':<20}{genders[index]:<30}")
            print(f"{'Guardian_Name':<20}{guardian_names[index]:<30}")
            print(f"{'Guardian_Number':<20}{guardian_numbers[index]:<30}")
            print(f"{'Date_of_Birth':<20}{birth_dates[index]:<30}")
            print(f"{'Registration_Date':<20}{registration_dates[index]:<30}")
            print(f"{'Class':<20}{classes[index]:<30}")
            print(f"{'Monthly Exam 1':<20}{monthly_exam_1[index -1]:<30}")
            print(f"{'Mid Term':<20}{mid_term[index-1]:<30}")
            print(f"{'Monthly Exam 2':<20}{monthly_exam_2[index-1]:<30}")
            print(f"{'Final Exam':<20}{final_exam[index-1]:<30}")
            print(f"{'Total Marks':<20}{total_marks[index-1]:<30}")
            print(f"{'-'*50}\n")
        else:
            print('Ma Diiwan_gelsano Id-kaan ee Ku noqo')
        
        # Ask if the user wants to continue
        per = input('Ma rabtaa inaa sii wado [1.Haa/2.Maya]: ')
        if per == '1':
            continue
        elif per == '2':
            c_s = False
        else:
            print('Soo Gali number-kaliya')
def att():
    
    print('1.Form One\n2.Form Two\n3.Form Three\n4.Form Four')
    print('')

    while True:
        
        bb = input('Fasalka : ')
        Fasal = None

        
        if bb == '1':
            Fasal = 'Form1'
        elif bb == '2':
            Fasal = 'Form2'
        elif bb == '3':
            Fasal = 'Form3'
        elif bb == '4':
            Fasal = 'Form4'
        else:
            print('Fadlan ku celi mar kale')
            continue
        xy = pd.read_excel('Attendence.xlsx', sheet_name=Fasal) 
        iD = xy.ID.tolist()
        name = xy.Magac_saddexan.tolist()

        if not iD:
            print('Fasalkan cidna kama diiwan-gelshano')
            continue
        attendance = []
        for i in range(len(name)):
            print('Enter attendance (P for Present, A for Absent):')
            while True:
                er = input(f'{iD[i]} {name[i]} : ')
                if er.capitalize() == 'P':
                    attendance.append(0) 
                    break
                elif er.capitalize() == 'A':
                    attendance.append(1)  
                    break
                else:
                    print('Soo gali kaliya A or P')
        xy['Absent'] = xy['Absent'] + attendance 
        with pd.ExcelWriter('Attendence.xlsx',mode='a',engine='openpyxl',if_sheet_exists='overlay')as pp:
            xy.to_excel(pp,sheet_name=Fasal,index=False)
        

        print('Attendance recorded:', attendance)
        cont = input('Ma rabtaa inaad sii wado? [Haa/Maya]: ')
        if cont.capitalize() == 'Maya':
            break
def Get_Exam_result():
    exam_data = pd.read_excel('Exam result.xlsx')
    student_ids = exam_data.ID_Number.tolist()
    full_names = exam_data.Name.tolist()

    monthly_exam_1 = exam_data.Monthly_Exam_1.tolist()
    mid_term = exam_data.Mid_term.tolist()
    monthly_exam_2 = exam_data.Monthly_Exam_2.tolist()
    final_exam = exam_data.Final.tolist()
    total_marks = exam_data.Total.tolist()

    while True:
        student_id = input('Enter the ID of the student: ').upper()
        if student_id in student_ids:
            index = student_ids.index(student_id)  # Find the index of the ID

            # Display student information
            print("--- Student Exam Result ---")
            print('')
            print(f"{'Field':<20}{'Value':<30}")
            print(f"{'-'*50}")
            print(f"{'Student ID':<20}{student_id:<30}")
            print(f"{'Name':<20}{full_names[index]:<30}")
            print(f"{'Monthly Exam 1':<20}{monthly_exam_1[index - 1]:<30}")
            print(f"{'Mid Term':<20}{mid_term[index -1]:<30}")
            print(f"{'Monthly Exam 2':<20}{monthly_exam_2[index-1]:<30}")
            print(f"{'Final Exam':<20}{final_exam[index-1]:<30}")
            print(f"{'Total Marks':<20}{total_marks[index-1]:<30}")
            print(f"{'-'*50}\n")
            break
        else:
            print('This ID is not registered. Try again.')
        
        # Ask if the user wants to continue
    while True:
        repeat = input("Do you want to get another student's exam result (1. Yess | 2. No): ")
        if repeat == '1':
            Get_Exam_result()
        elif repeat == '2':
            exit()
        else:
            print('Enter 1 or 2')
def registeration():
    # Load the data
    main_data = pd.read_excel('Teacher registeration.xlsx')
    Teacher_ID = main_data.ID.tolist()
    Teacher_Number = main_data.Number.tolist() 
    Teacher_Name = main_data.Name.tolist()
    degrees = [
    # Bachelor's Degrees
    "Bachelor Of Computer Science",
    "Bachelor Of Psychology",
    "Bachelor Of Business Administration",
    "Bachelor Of Civil Engineering",
    "Bachelor Of Medicine and Surgery",
    "Bachelor Of Medicine",
    "Bachelor Of Fine Arts",
    "Bachelor Of Environmental Science",
    "Bachelor Of Biotechnology",
    "Bachelor Of Sociology",
    "Bachelor Of Nursing",
    "Bachelor Of Mathematics",
    "Bachelor Of Physics",
    "Bachelor Of Chemistry",
    "Bachelor Of Political Science",
    "Bachelor Of Geology",
    "Bachelor Of History",
    "Bachelor Of English Literature",
    "Bachelor Of Zoology",
    "Bachelor Of Anthropology",
    "Bachelor Of Education",
    "Bachelor Of Agriculture",
    "Bachelor Of Journalism and Mass Communication",
    "Bachelor Of Statistics",
    "Bachelor Of Electronics",
    "Bachelor Of Philosophy",
    "Bachelor Of Visual Arts",
    "Bachelor Of Microbiology",
    "Bachelor Of Genetics",
    "Bachelor Of Marine Biology",
    "Bachelor Of Foreign Languages",
    "Bachelor Of Architecture",
    "Bachelor Of Fashion Design",
    "Bachelor Of Computer Engineering",
    "Bachelor Of Information Technology",
    "Bachelor Of Commerce",
    "Bachelor Of Laws",
    "Bachelor Of Forensic Science",
    "Bachelor Of Public Administration",
    "Bachelor Of Nutrition and Dietetics",
    "Bachelor Of Environmental Engineering",
    "Bachelor Of Robotics",
    "Bachelor Of Astronomy",
    "Bachelor Of Cybersecurity",
    "Bachelor Of Cloud Computing",
    "Bachelor Of Artificial Intelligence",
    "Bachelor Of Business Management",
    "Bachelor Of Tourism and Hospitality Management",
    "Bachelor Of Social Work",
    "Bachelor Of Data Analytics",
    "Bachelor Of Media Studies",
    "Bachelor Of Aerospace Engineering",
    "Bachelor Of Software Engineering",
    "Bachelor Of Theatre and Drama",
    "Bachelor Of Music",
    "Bachelor Of Renewable Energy",
    "Bachelor Of Supply Chain Management",
    "Bachelor Of Criminology",
    "Bachelor Of Sculpture",
    "Bachelor Of Digital Marketing",
    "Bachelor Of Health Informatics",
    "Bachelor Of Electronics and Communication Engineering",
    "Bachelor Of Mechanical Engineering",
    "Bachelor Of Biomedical Engineering",
    "Bachelor Of International Relations",
    "Bachelor Of Development Studies",
    "Bachelor Of Applied Physics",
    "Bachelor Of Cognitive Science",
    "Bachelor Of Oceanography",
    "Bachelor Of Wildlife Conservation",
    "Bachelor Of Film Studies",
    "Bachelor Of Game Development",
    "Bachelor Of Petroleum Engineering",
    "Bachelor Of Textile Engineering",
    "Bachelor Of Peace and Conflict Studies",
    "Bachelor Of Structural Engineering",
    "Bachelor Of Gender Studies",
    "Bachelor Of Civil Engineering",
    "Bachelor Of Classical Studies",
    "Bachelor Of Urban Studies",
    "Bachelor Of Nanotechnology",
    "Bachelor Of Music Therapy",
    "Bachelor Of Human Genetics",
    "Bachelor Of Environmental Health",
    "Bachelor Of Energy Engineering",
    "Bachelor Of Industrial Design",
    "Bachelor Of Metallurgical Engineering",
    "Bachelor Of Agricultural Engineering",
    "Bachelor Of Cultural Studies",
    "Bachelor Of Disaster Management",
    "Bachelor Of Game Design",
    "Bachelor Of Instrumentation Engineering",
    "Bachelor Of Sustainable Development",
    "Bachelor Of Adventure Tourism",
    "Bachelor Of Scriptwriting",
    "Bachelor Of Agronomy",
    "Bachelor Of Wildlife Biology",
    "Bachelor Of Computational Biology",
    "Bachelor Of Human Rights",
    "Bachelor Of Applied Mathematics",
    "Bachelor Of Visual Communication",
    
    # Master's Degrees
    "Master Of Artificial Intelligence",
    "Master Of International Relations",
    "Master Of Business Administration",
    "Master Of Public Health",
    "Master Of Robotics",
    "Master Of Creative Writing",
    "Master Of Cybersecurity",
    "Master Of Data Science",
    "Master Of Renewable Energy",
    "Master Of Journalism",
  
    # PhD Degrees
    "PhD Of Computer Science",
    "PhD Of Artificial Intelligence",
    "PhD Of Data Science",
    "PhD Of Environmental Science",
    "PhD Of Physics",
    "PhD Of Chemistry",
    "PhD Of Biology",
    "PhD Of Biotechnology",
    "PhD Of Mathematics",
    "PhD Of Mechanical Engineering",
    ]
    ma_r = True

    
    def generate_next_id(existing_ids):
            if not existing_ids:
                return 'MT001'
            max_id = max(int(i[2:]) for i in existing_ids if i.startswith('MT'))
            return f"MT{max_id + 1:03}" 
    # print(generate_next_id(Teacher_ID))



    while True:
        try:
            Number_of_Teachers=int(input("How many teachers you want to register: ".title()))
        except ValueError:
            print("Enter Number Only")
            continue

        for i in range(1,Number_of_Teachers+1):
            # Automatically generate the ID
            iD=generate_next_id(Teacher_ID)
            
            # Validate The Full Name
            while True:
                Name=input(f"Enter the full name of the teacher (Name Father_Name Grand_Father_Name): ".title()).title().strip().split()
                if Name in Teacher_Name:  
                     print('This name is already registered'.title())
                elif all(len(i) >= 3 for i in Name) and len(Name) == 3:
                    break
                else:
                    print("Enter the Full Name and Each Name Must contain at least 3 Letters")
                print("")

            # Validate Number
            while True:
                Number = input("Enter The Teacher's Number (061xxxxxxx): ")
                if Number in Teacher_Number :
                    print('This number is already registered'.title())
                elif Number.startswith('061') and len(Number) == 10 and Number.isdigit():
                    break
                else:
                    print("Wrong Number.")

            # Validate Sex
            while True:
                Sex = input("Enter Teacher's Sex (Male or Female): ").capitalize()
                if Sex in ['Male', 'Female']:
                    break
                else:
                    print("Enter Male or Female")

            # Validate Degree
            while True:
                Degree = input("Enter Teacher's Degree: ").title()
                print(Degree)
                if Degree in degrees:
                    break
                else:
                    print("Enter Valid Degree")
            
            # Validate Degree
            while True:
                Subject = input("Enter Teacher's Subject: ").title()
                if Subject in ['Math','Biology','Chemistry','Physics','Somali','English','ArabiC','Geography','History','Technoogy','Business','Islamic Education']:
                    print(f"{iD} {" ".join(Name)} is successfully registered")
                    print(f"The ID of Teacher {" ".join(Name)} is: {iD}")
                    break
                else:
                    print("Enter Valid Subject")

            # Saving to excel
            new_row = {
                'ID': iD,
                'Name': " ".join(Name),
                'Number': Number,
                'Sex': Sex,
                'Degree': Degree,
                'Subject':Subject,
                'Date': time.strftime('%d/%m/20%y')
            }
            main_data = pd.concat([main_data, pd.DataFrame([new_row])], ignore_index=True)

            # Save the updated data
            with pd.ExcelWriter('Teacher registeration.xlsx', mode='a', engine='openpyxl',if_sheet_exists = 'overlay') as ui:
                main_data.to_excel(ui, index=False)
            Teacher_ID.append(iD)
            Teacher_Name.append(Name)
            Teacher_Number.append(Number)
         
        break
def Teacher_Data_Search():
    # Load data from Excel files
    Teacher_regestration_data = pd.read_excel('Teacher registeration.xlsx', dtype={'Number': str})
    # exam_data = pd.read_excel('Exam result.xlsx')
    
    # Extract the necessary columns into lists
    Teacher_ids = Teacher_regestration_data.ID.tolist()
    full_names = Teacher_regestration_data.Name.tolist()
    phone_numbers = Teacher_regestration_data.Number.tolist()
    Sexs = Teacher_regestration_data.Sex.tolist()
    Teachers_Degrees = Teacher_regestration_data.Degree.tolist()
    Subjects = Teacher_regestration_data.Subject.tolist()
    registration_dates = Teacher_regestration_data.Date.tolist()

    while True:
        Teacher_ID = input('Enter the ID of the Teacher: ').upper()
        if Teacher_ID in Teacher_ids:
            index = Teacher_ids.index(Teacher_ID)  # Find the index of the ID

            # Display student information
            print("\n--- Teacher Data ---")
            print('')
            print(f"{'Field':<20}{'Value':<30}")
            print(f"{'-'*50}")
            print(f"{'ID':<20}{Teacher_ID:<30}")
            print(f"{'Name':<20}{full_names[index]:<30}")
            print(f"{'Number':<20}{phone_numbers[index]:<30}")
            print(f"{'Sex':<20}{Sexs[index]:<30}")
            print(f"{'Degree':<20}{Teachers_Degrees[index]:<30}")
            print(f"{'Subject':<20}{Subjects[index]:<30}")
            print(f"{'Date':<20}{registration_dates[index]:<30}")
            print(f"{'-'*50}\n")
            break
        else:
            print('This ID is not registered. Try again.')
        
        # Ask if the user wants to continue
    while True:
        per = input("Do you want to get another teacher's data: ")
        if per == '1':
            continue
        elif per == '2':
            break
        else:
            print('Enter number only')
def F1(Fasalka):
    print(f" \
          \nJadwalka Form one (F{Fasalka})")
    print('''----------------------------------------------------------------------------------------------------------
Maalin         | Xiisada 1  | Xiisada 2  | Xiisada 3  | Xiisada 4  | Xiisada 5  | Xiisada 6  | Xiisada 7
----------------------------------------------------------------------------------------------------------
Sabti          | Tarbiyo    | Bayooloji  | Ingiriisi  | Soomaali   | Taariikh   | Carabi     | Kimistar  
Axad           | Taariikh   | Ganacsi    | Tarbiyo    | Fiisigis   | Ingiriisi  | Juqraafi   | Bayooloji 
Isniin         | Xisaab     | Ingiriisi  | Carabi     | Ganacsi    | Fiisigis   | Juqraafi   | Soomaali  
Talaado        | Carabi     | Juqraafi   | Tiknoolaji | Ganacsi    | Fiisigis   | Tarbiyo    | Kimistar  
Arbaco         | Xisaab     | Juqraafi   | Taariikh   | Soomaali   | Ingiriisi  | Carabi     | Tiknoolaji
----------------------------------------------------------------------------------------------------------''')
def F2(Fasalka):
    print(f" \
          \nJadwalka Form Two (F{Fasalka})")
    print('''-----------------------------------------------------------------------------------------------------------
Maalin         | Xiisada 1  | Xiisada 2  | Xiisada 3  | Xiisada    | Xiisada 5  | Xiisada 6  | Xiisada 7
-----------------------------------------------------------------------------------------------------------
Sabti          | Kimistar   | Taariikh   | Carabi     | Soomaali   | Fiisigis   | Tarbiyo    | Bayooloji
Axad           | Juqraafi   | Tarbiyo    | Tiknoolaji | Ganacsi    | Taariikh   | Fiisigis   | Kimistar
Isniin         | Fiisigis   | Ingiriisi  | Xisaab     | Juqraafi   | Kimistar   | Bayooloji  | Tiknoolaji
Talaado        | Ingiriisi  | Carabi     | Tiknoolaji | Kimistar   | Soomaali   | Bayooloji  | Fiisigis
Arbaco         | Tarbiyo    | Taariikh   | Carabi     | Ingiriisi  | Bayooloji  | Juqraafi   | Soomaali
-----------------------------------------------------------------------------------------------------------''')
def F3(Fasalka):
    print(f" \
          \nJadwalka Form Three (F{Fasalka})")
    print('''---------------------------------------------------------------------------------------------------------
Maalin         | Xiisada 1  | Xiisada 2  | Xiisada 3  | Xiisada 4  | Xiisada 5  | Xiisada 6  | Xiisada 7
---------------------------------------------------------------------------------------------------------
Sabti          | Taariikh   | Juqraafi   | Kimistar   | Tiknoolaji | Ganacsi    | Bayooloji  | Ingiriisi
Axad           | Bayooloji  | Soomaali   | Ganacsi    | Carabi     | Fiisigis   | Xisaab     | Ingiriisi
Isniin         | Tarbiyo    | Tiknoolaji | Soomaali   | Bayooloji  | Fiisigis   | Taariikh   | Ingiriisi
Talaado        | Taariikh   | Kimistar   | Ingiriisi  | Soomaali   | Juqraafi   | Ganacsi    | Carabi
Arbaco         | Juqraafi   | Taariikh   | Fiisigis   | Xisaab     | Ingiriisi  | Carabi     | Kimistar
---------------------------------------------------------------------------------------------------------''')
def F4(Fasalka):
    print(f" \
          \nJadwalka Form Four (F{Fasalka})")
    print('''--------------------------------------------------------------------------------------------------------
Maalin         | Xiisada 1  | Xiisada 2  | Xiisada 3  | Xiisada 4  | Xiisada 5  | Xiisada 6  | Xiisada 7
--------------------------------------------------------------------------------------------------------
Sabti          | Soomaali   | Taariikh   | Tarbiyo    | Ganacsi    | Carabi     | Kimistar   | Fiisigis
Axad           | Ingiriisi  | Tarbiyo    | Juqraafi   | Soomaali   | Fiisigis   | Carabi     | Kimistar
Isniin         | Bayooloji  | Xisaab     | Ingiriisi  | Kimistar   | Taariikh   | Tiknoolaji | Juqraafi
Talaado        | Taariikh   | Xisaab     | Tiknoolaji | Tarbiyo    | Carabi     | Ingiriisi  | Soomaali
Arbaco         | Carabi     | Ingiriisi  | Soomaali   | Tiknoolaji | Ganacsi    | Juqraafi   | Xisaab
--------------------------------------------------------------------------------------------------------''')
def timetable():
    print("1.Form One (F1)\
          \n2.Form Two (F2)\
          \n3.Form Three (F3)\
          \n4.Form Four (F4) ")

    while True:
        try:
            Fasalka = int(input("Soo gali fasalka aad rabtid: ".title()))
            if Fasalka == 1:
                F1(Fasalka)
            elif Fasalka == 2:
                F2(Fasalka)
            elif Fasalka == 3:
                F3(Fasalka)
            elif Fasalka == 4:
                F4(Fasalka)
            else:
                print("Majiro")
            break
        except ValueError:
            print("Lambar soo gali")
def set_exam_result():
    import pandas as pd
    from openpyxl import load_workbook

    # Load workbook and data
    workbook = load_workbook("Exam result.xlsx")
    xy = pd.read_excel('Exam result.xlsx')
    S_ID = xy.ID_Number.tolist()
    Monthly_Exam_1_Result = []
    Mid_term_Result = []
    Monthly_Exam_2_Result = []
    Final_Result = []
    Name = []
    total = []

    Subjects = []
    for i in range(1, 13):
        Subjects.append(i)

    ID = []
    result = []

    print("1.Math \n2.Biology \n3.Chemistry \n4.Physics \n5.Somali \n6.English \n7.Arabic \n8.Geography \n9.History \n10.Technology \n11. Business \n12.Islamic Education")
    while True:
        try:
            Subject = int(input("Enter the subject: ".title()))
            if Subject in Subjects:
                print("1.Form 1 \n2.Form 2 \n3.Form 3 \n4.Form 4")
                while True:
                    try:
                        Class = int(input("Enter the class: ".title()))
                        if Class in [1, 2, 3, 4]:
                            while True:
                                try:
                                    Number_students = int(input("Enter the number of students: ".title()))
                                    for i in range(1, Number_students + 1):
                                        while True:
                                            Student_ID = input(f"Enter the id of student {i}: ".title())
                                            if Student_ID in S_ID:
                                                ID.append(Student_ID)
                                                Name.append(None)
                                                total.append(None)
                                                Monthly_Exam_1_Result.append(None)
                                                Mid_term_Result.append(None)
                                                Monthly_Exam_2_Result.append(None)
                                                Final_Result.append(None)
                                                break
                                            else:
                                                print("This ID is not registered. Try again.")

                                        while True:
                                            try:
                                                student_result = int(input(f"Enter the result of student {i}: ".title()))
                                                if student_result in Result:
                                                    result.append(student_result)
                                                    break
                                                else:
                                                    print("Wrong Result")
                                            except:
                                                print("Enter Number Only!")
                                    break
                                except:
                                    print("Enter Number Only!")
                            break
                        else:
                            print("Unknown Class")
                    except:
                        print("Enter Number Only")
                break
            else:
                print("Unknown Subject")
        except:
            print("Enter Number Only")


    def Monthly_Exam_1():
        global Result, Monthly_Exam_1_Result
        Result = []
        for i in range(21):
            Result.append(i)
        set_exam_result()


    def Mid_term():
        global Result, Mid_term_Result
        Result = []
        for i in range(31):
            Result.append(i)
        set_exam_result()


    def Monthly_Exam_2():
        global Result, Monthly_Exam_2_Result
        Result = []
        for i in range(21):
            Result.append(i)
        set_exam_result()


    def Final():
        global Result, Final_Result
        Result = []
        for i in range(31):
            Result.append(i)
        set_exam_result()


    def Saving_excel():
        file = 'Exam result.xlsx'
        sheet_name = f"Form{Class}"

        df = pd.read_excel(file, sheet_name=sheet_name)
        data = {
            'ID_Number': ID,
            'Name': Name,
            'Monthly_Exam_1': Monthly_Exam_1_Result,
            'Mid_term': Mid_term_Result,
            'Monthly_Exam_2': Monthly_Exam_2_Result,
            'Final': Final_Result,
            'Total': total
        }
        df1 = pd.DataFrame(data)
        for i, row in df1.iterrows():
            existing_row = df[df['ID_Number'] == row['ID_Number']]
            if not existing_row.empty:
                for col in ['Monthly_Exam_1', 'Mid_term', 'Monthly_Exam_2', 'Final']:
                    if pd.isna(existing_row.iloc[0][col]) and not pd.isna(row[col]):
                        df.loc[existing_row.index, col] = row[col]
            else:
                df = pd.concat([df, row.to_frame().T], ignore_index=True)

        with pd.ExcelWriter(file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)


    # Menu for selecting exam type
    print("1.Monthly Exam 1\n2.Mid-Term\n3.Monthly Exam 2\n4.Final\n5.Exit")
    while True:
        try:
            Exam_Type = int(input("Enter the Exam_Type you want: ".title()))
            if Exam_Type == 1:
                Monthly_Exam_1()
            elif Exam_Type == 2:
                Mid_term()
            elif Exam_Type == 3:
                Monthly_Exam_2()
            elif Exam_Type == 4:
                Final()
            elif Exam_Type == 5:
                print("Exiting exam result management system.".title())
                break
            else:
                print("Invalid choice!".title())
        except:
            print("Enter Number Only!")

    # Update results based on the exam type
    if Exam_Type == 1:
        Monthly_Exam_1_Result.clear()
        Monthly_Exam_1_Result.extend(result)
    elif Exam_Type == 2:
        Mid_term_Result.clear()
        Mid_term_Result.extend(result)
    elif Exam_Type == 3:
        Monthly_Exam_2_Result.clear()
        Monthly_Exam_2_Result.extend(result)
    elif Exam_Type == 4:
        Final_Result.clear()
        Final_Result.extend(result)

    Saving_excel()

def login():
    print(f'\n{" " * 20}--- Al-Anwaar Primary & Secondary School ---')
    print('')
    rt = True
    while rt:
        username = input('Enter username: ')
        password = input('Enter password : ')
        print('\n')
        if username =='Mahad_123' and password == 'B5Sc1528':
            print('1.Student Management Sytem\n2.Teacher Management System\n3.Timetable Management System\n4.Exam Result Management System')
            print('')
            l = True
            while l:
                ty = input('Enter one of them : ')
                if ty =='1':
                    print('1.Register Student \n2.Record attendance\n3.Get Student Data') 
                    while True:
                        tt = input('Enter one of them : ')
                        if tt == '1':
                            register_students()
                            l = False
                            rt = False
                            break
                        elif tt =='2':
                            att()
                            l = False
                            rt = False
                            break
                        elif tt =='3':
                            Xog_raadin()
                            l = False
                            rt = False
                            break
                        else:
                            print('Invalid') 

                elif ty == '2':
                    print('1.Register Teacher\n2.Get Teacher Data')
                    while True:
                        tt = input('Enter one of them : ')
                        if tt =='1':
                            registeration()
                            l = False
                            rt = False
                            break
                        elif tt == '2':
                            Teacher_Data_Search()
                            l = False
                            rt = False
                            break
                        else:
                            print('Invalid')
                elif ty == '3':
                    timetable()
                    l = False
                    rt = False
                elif ty == '4':
                    print('1.Record Exam result\n2.Get Exam result')
                    while True:
                        tt = input('Enter one of them : ')
                        if tt == '1':
                            set_exam_result
                            l = False
                            rt = False
                            break
                        elif tt == '2':
                            Get_Exam_result()
                            l = False
                            rt = False
                            break
                        else:
                            print('Invalid')
                        
                else:
                    print('invalid') 

        else:
            print('Username and password are invalid') 
    
    
login()
